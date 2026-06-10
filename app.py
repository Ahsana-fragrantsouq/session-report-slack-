import os
import io
import requests
import openpyxl
from datetime import datetime, timedelta
from zoneinfo import ZoneInfo
from flask import Flask, jsonify
from apscheduler.schedulers.background import BackgroundScheduler

app = Flask(__name__)

# ── ENV VARS ──────────────────────────────────────────────────────────────────
SHOPIFY_STORE        = os.environ.get("SHOPIFY_STORE", "fragrantsouq.myshopify.com")
SHOPIFY_ACCESS_TOKEN = os.environ.get("SHOPIFY_ADMIN_API_TOKEN") or os.environ.get("SHOPIFY_ACCESS_TOKEN")  # shpat_...
SLACK_BOT_TOKEN      = os.environ.get("SLACK_BOT_TOKEN")        # xoxb-...
SLACK_CHANNEL_ID     = os.environ.get("SLACK_CHANNEL_ID", "C0B9V9U312L")

IST = ZoneInfo("Asia/Kolkata")


# ── SHOPIFY: fetch sessions by landing page ───────────────────────────────────
def fetch_sessions(date_str):
    print(f"[fetch_sessions] Querying Shopify Admin API for date: {date_str}", flush=True)

    url = f"https://{SHOPIFY_STORE}/admin/api/2026-04/graphql.json"
    print(f"[fetch_sessions] URL: {url}", flush=True)

    headers = {
        "Content-Type": "application/json",
        "X-Shopify-Access-Token": SHOPIFY_ACCESS_TOKEN,
    }

    # First introspect to find correct field names on ShopifyqlTableData
    introspect = """
    {
      __type(name: "ShopifyqlTableData") {
        fields { name }
      }
    }
    """
    ir = requests.post(url, json={"query": introspect}, headers=headers, timeout=30)
    idata = ir.json()
    fields = [f["name"] for f in idata.get("data", {}).get("__type", {}).get("fields", [])]
    print(f"[fetch_sessions] ShopifyqlTableData fields: {fields}", flush=True)

    # Introspect rows field type on ShopifyqlTableData
    introspect2 = """
    {
      __type(name: "ShopifyqlTableData") {
        fields {
          name
          type {
            name
            kind
            ofType { name kind ofType { name kind } }
          }
        }
      }
    }
    """
    try:
        ir2 = requests.post(url, json={"query": introspect2}, headers=headers, timeout=30)
        print(f"[fetch_sessions] Introspect2 status: {ir2.status_code}", flush=True)
        idata2 = ir2.json()
        print(f"[fetch_sessions] Introspect2 raw: {idata2}", flush=True)
        tdata_fields = idata2.get("data", {}).get("__type", {}).get("fields", [])
        for f in tdata_fields:
            print(f"[fetch_sessions] Field '{f['name']}' type: {f['type']}", flush=True)
    except Exception as e2:
        print(f"[fetch_sessions] Introspect2 error: {e2}", flush=True)

    query = """
    {
      shopifyqlQuery(query: "FROM sessions SHOW landing_page_type, landing_page_path, online_store_visitors, sessions SINCE -1d UNTIL -1d ORDER BY sessions DESC") {
        tableData {
          columns { name }
          rows
        }
        parseErrors
      }
    }
    """

    print(f"[fetch_sessions] Sending ShopifyQL request...", flush=True)
    resp = requests.post(url, json={"query": query}, headers=headers, timeout=30)
    print(f"[fetch_sessions] Response status: {resp.status_code}", flush=True)
    data = resp.json()
    print(f"[fetch_sessions] Raw response keys: {list(data.keys())}", flush=True)

    if "errors" in data:
        print(f"[fetch_sessions] GraphQL errors: {data['errors']}", flush=True)
        raise RuntimeError(f"GraphQL errors: {data['errors']}")

    shopify_data = data.get("data", {}).get("shopifyqlQuery", {})
    parse_errors = shopify_data.get("parseErrors")
    if parse_errors:
        print(f"[fetch_sessions] ShopifyQL parse errors: {parse_errors}", flush=True)
        raise ValueError(f"ShopifyQL parse errors: {parse_errors}")

    table = shopify_data.get("tableData", {})
    if not table:
        print(f"[fetch_sessions] WARNING - No tableData returned from Shopify.", flush=True)
        return []

    columns = [col["name"] for col in table.get("columns", [])]
    rows    = table.get("rows", [])
    print(f"[fetch_sessions] Columns: {columns}", flush=True)
    print(f"[fetch_sessions] Total rows: {len(rows)}", flush=True)
    if rows:
        print(f"[fetch_sessions] First row sample: {rows[0]}", flush=True)

    # rows is a JSON scalar — each row is a list of values matching columns order
    results = []
    for row in rows:
        if isinstance(row, list):
            record = dict(zip(columns, row))
        elif isinstance(row, dict):
            record = row
        else:
            continue
        results.append(record)

    print(f"[fetch_sessions] Successfully parsed {len(results)} rows.", flush=True)
    return results


# ── BUILD EXCEL ───────────────────────────────────────────────────────────────
def build_excel(rows, date_str):
    """
    Filters Product rows, keeps 3 columns, returns Excel bytes.
    """
    print(f"[build_excel] Starting Excel build for {date_str}. Input rows: {len(rows)}", flush=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sessions by Landing Page"

    # Header
    headers = ["Landing page path", "Online store visitors", "Sessions"]
    ws.append(headers)

    # Style header row
    from openpyxl.styles import Font, PatternFill, Alignment
    header_fill = PatternFill("solid", fgColor="1F4E79")
    for cell in ws[1]:
        cell.font      = Font(bold=True, color="FFFFFF")
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Data rows — only Product type
    count = 0
    skipped = 0
    for r in rows:
        page_type = str(r.get("landing_page_type", "")).strip().lower()
        if page_type != "product":
            print(f"[build_excel] Skipping row — type='{page_type}', path='{r.get('landing_page_path', '')}'", flush=True)
            skipped += 1
            continue
        ws.append([
            r.get("landing_page_path", ""),
            r.get("online_store_visitors", 0),
            r.get("sessions", 0),
        ])
        count += 1

    print(f"[build_excel] Rows added (Product): {count} | Rows skipped (non-Product): {skipped}", flush=True)

    # Column widths
    ws.column_dimensions["A"].width = 70
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 12

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    excel_size_kb = round(len(buf.getvalue()) / 1024, 2)
    print(f"[build_excel] Excel file built successfully. Size: {excel_size_kb} KB", flush=True)
    return buf.getvalue(), count


# ── SEND TO SLACK ─────────────────────────────────────────────────────────────
def send_to_slack(excel_bytes, date_str, row_count):
    filename = f"sessions_by_landing_page_{date_str}.xlsx"
    file_size = len(excel_bytes)
    print(f"[send_to_slack] Preparing to upload '{filename}' to Slack channel {SLACK_CHANNEL_ID}", flush=True)
    print(f"[send_to_slack] File size: {round(file_size/1024, 2)} KB | Product rows: {row_count}", flush=True)

    # Step 1 — Get upload URL
    print(f"[send_to_slack] Step 1: Requesting upload URL from Slack...", flush=True)
    url_resp = requests.post(
        "https://slack.com/api/files.getUploadURLExternal",
        headers={"Authorization": f"Bearer {SLACK_BOT_TOKEN}"},
        data={"filename": filename, "length": file_size},
        timeout=30,
    )
    url_resp.raise_for_status()
    url_result = url_resp.json()
    print(f"[send_to_slack] Step 1 response: ok={url_result.get('ok')} error={url_result.get('error','none')}", flush=True)

    if not url_result.get("ok"):
        raise RuntimeError(f"Slack getUploadURLExternal failed: {url_result.get('error')}")

    upload_url = url_result["upload_url"]
    file_id    = url_result["file_id"]
    print(f"[send_to_slack] Got upload URL. File ID: {file_id}", flush=True)

    # Step 2 — Upload file bytes to the upload URL
    print(f"[send_to_slack] Step 2: Uploading file bytes...", flush=True)
    upload_resp = requests.post(
        upload_url,
        files={"file": (filename, excel_bytes,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        timeout=60,
    )
    upload_resp.raise_for_status()
    print(f"[send_to_slack] Step 2 done. HTTP status: {upload_resp.status_code}", flush=True)

    # Step 3 — Complete upload and share to channel
    print(f"[send_to_slack] Step 3: Completing upload and posting to channel...", flush=True)
    complete_resp = requests.post(
        "https://slack.com/api/files.completeUploadExternal",
        headers={"Authorization": f"Bearer {SLACK_BOT_TOKEN}"},
        json={
            "files": [{"id": file_id, "title": filename}],
            "channel_id": SLACK_CHANNEL_ID,
            "initial_comment": (
                f":bar_chart: *Sessions by Landing Page — {date_str}*\n"
                f"Product pages: *{row_count} rows* | Columns: Landing page path, Online store visitors, Sessions"
            ),
        },
        timeout=30,
    )
    complete_resp.raise_for_status()
    complete_result = complete_resp.json()
    print(f"[send_to_slack] Step 3 response: ok={complete_result.get('ok')} error={complete_result.get('error','none')}", flush=True)

    if not complete_result.get("ok"):
        raise RuntimeError(f"Slack completeUploadExternal failed: {complete_result.get('error')}")

    print(f"[send_to_slack] SUCCESS - File posted to Slack channel. File ID: {file_id}", flush=True)
    print(f"[session-slack] Sent {row_count} product rows for {date_str} to Slack.", flush=True)


# ── MAIN JOB ──────────────────────────────────────────────────────────────────
def run_job():
    print(f"[run_job] ─────────────────────────────────────────", flush=True)
    print(f"[run_job] Job started at {datetime.now(IST).strftime('%Y-%m-%d %H:%M:%S IST')}", flush=True)
    try:
        yesterday = (datetime.now(IST) - timedelta(days=1)).strftime("%Y-%m-%d")
        print(f"[run_job] Target date (yesterday): {yesterday}", flush=True)

        rows = fetch_sessions(yesterday)
        print(f"[run_job] Got {len(rows)} total rows from Shopify.", flush=True)

        excel_bytes, count = build_excel(rows, yesterday)
        print(f"[run_job] Filtered to {count} Product rows.", flush=True)

        send_to_slack(excel_bytes, yesterday, count)
        print(f"[run_job] Job completed successfully.", flush=True)

    except Exception as e:
        print(f"[run_job] ERROR: {e}", flush=True)
    print(f"[run_job] ─────────────────────────────────────────", flush=True)


# ── FLASK ROUTES ──────────────────────────────────────────────────────────────
@app.route("/health")
def health():
    print(f"[health] Health check called at {datetime.now(IST).strftime('%Y-%m-%d %H:%M:%S IST')}", flush=True)
    return jsonify({"status": "ok"})


@app.route("/run-now")
def run_now():
    """Trigger the job manually for testing."""
    print(f"[run-now] Manual trigger called at {datetime.now(IST).strftime('%Y-%m-%d %H:%M:%S IST')}", flush=True)
    run_job()
    return jsonify({"status": "done"})


@app.route("/run-date/<date_str>")
def run_date(date_str):
    """Fetch a specific date. Format: YYYY-MM-DD  e.g. /run-date/2026-06-07"""
    print(f"[run-date] Manual date trigger called for: {date_str}", flush=True)
    try:
        rows = fetch_sessions(date_str)
        excel_bytes, count = build_excel(rows, date_str)
        send_to_slack(excel_bytes, date_str, count)
        print(f"[run-date] Completed for {date_str}. Product rows: {count}", flush=True)
        return jsonify({"status": "done", "product_rows": count, "date": date_str})
    except Exception as e:
        print(f"[run-date] ERROR for {date_str}: {e}", flush=True)
        return jsonify({"status": "error", "message": str(e)}), 500


# ── SCHEDULER ─────────────────────────────────────────────────────────────────
scheduler = BackgroundScheduler(timezone=IST)
scheduler.add_job(run_job, "cron", hour=9, minute=0)
scheduler.start()
print(f"[startup] Scheduler started. Job will run daily at 09:00 IST.", flush=True)
print(f"[startup] SHOPIFY_STORE     : {SHOPIFY_STORE}", flush=True)
print(f"[startup] SHOPIFY_TOKEN set : {'YES' if SHOPIFY_ACCESS_TOKEN else 'NO ⚠️'}", flush=True)
print(f"[startup] SLACK_TOKEN set   : {'YES' if SLACK_BOT_TOKEN else 'NO ⚠️'}", flush=True)
print(f"[startup] SLACK_CHANNEL_ID  : {SLACK_CHANNEL_ID}", flush=True)

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000)
